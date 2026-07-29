import csv
import logging
from typing import List
import openpyxl

logger = logging.getLogger(__name__)


class ExcelService:
    """Service for parsing Excel (.xlsx, .xls) and CSV (.csv) spreadsheets into structured, RAG-ready key-value text format."""

    def extract_text(self, file_path: str, file_type: str) -> str:
        """Extracts text from Excel or CSV file in structured row-by-row key-value format."""
        file_type = file_type.lower().strip()
        if file_type == "csv":
            return self._extract_csv(file_path)
        elif file_type in ["xlsx", "xls"]:
            return self._extract_excel(file_path)
        else:
            raise ValueError(f"Unsupported spreadsheet format: {file_type}")

    def _extract_csv(self, file_path: str) -> str:
        """Parses CSV file into structured key-value row representations."""
        formatted_rows: List[str] = []
        try:
            with open(file_path, mode="r", encoding="utf-8-sig", errors="replace") as f:
                reader = csv.reader(f)
                rows = list(reader)
                if not rows:
                    return ""

                headers = [h.strip() if h else f"Col_{i+1}" for i, h in enumerate(rows[0])]

                for idx, row in enumerate(rows[1:], start=2):
                    if not any(cell.strip() for cell in row if cell):
                        continue  # Skip completely empty rows

                    row_parts = [f"[Sheet: CSV Data | Row {idx}]"]
                    for col_idx, cell_value in enumerate(row):
                        header_name = headers[col_idx] if col_idx < len(headers) else f"Col_{col_idx+1}"
                        val_str = str(cell_value).strip()
                        if val_str:
                            row_parts.append(f"{header_name}: {val_str}")

                    if len(row_parts) > 1:
                        formatted_rows.append(" | ".join(row_parts))

        except Exception as e:
            logger.error(f"Error parsing CSV file {file_path}: {e}")
            raise RuntimeError(f"Failed to parse CSV document: {e}")

        return "\n\n".join(formatted_rows)

    def _extract_excel(self, file_path: str) -> str:
        """Parses Excel workbook (.xlsx/.xls) into structured key-value row representations with sheet metadata."""
        formatted_rows: List[str] = []
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_rows = list(sheet.iter_rows(values_only=True))
                if not sheet_rows:
                    continue

                # Find first non-empty row as header
                header_row_idx = 0
                headers = []
                for idx, row in enumerate(sheet_rows):
                    if any(cell is not None and str(cell).strip() != "" for cell in row):
                        header_row_idx = idx
                        headers = [str(c).strip() if c is not None and str(c).strip() != "" else f"Col_{i+1}" for i, c in enumerate(row)]
                        break

                if not headers:
                    continue

                # Forward-fill column state for categories
                last_seen_values = [None] * len(headers)

                for row_idx, row in enumerate(sheet_rows[header_row_idx + 1:], start=header_row_idx + 2):
                    if not any(cell is not None and str(cell).strip() != "" for cell in row):
                        continue  # Skip empty row

                    row_parts = [f"[Sheet: {sheet_name} | Row {row_idx}]"]
                    for col_idx, cell_value in enumerate(row):
                        if col_idx >= len(headers):
                            break
                        header_name = headers[col_idx]
                        if cell_value is not None and str(cell_value).strip() != "":
                            val_str = str(cell_value).strip()
                            last_seen_values[col_idx] = val_str
                            row_parts.append(f"{header_name}: {val_str}")
                        elif last_seen_values[col_idx] is not None:
                            # Forward-filled category fallback for merged structures
                            row_parts.append(f"{header_name}: {last_seen_values[col_idx]} (inherited)")

                    if len(row_parts) > 1:
                        formatted_rows.append(" | ".join(row_parts))

            wb.close()
        except Exception as e:
            logger.error(f"Error parsing Excel file {file_path}: {e}")
            raise RuntimeError(f"Failed to parse Excel document: {e}")

        return "\n\n".join(formatted_rows)
